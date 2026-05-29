#!/usr/bin/env python3
"""Local paper library manager implementation.

The public command remains ``scripts/paper_manager.py``. This module contains
the implementation so the CLI entrypoint can stay stable while internals are
split or refined over time.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


PACKAGE_DIR = Path(__file__).resolve().parent


def find_vault_root(start: Path) -> Path:
    """Find the Paper Library vault root from a script or package path."""
    for candidate in (start, *start.parents):
        if (candidate / ".obsidian").exists() and (candidate / "papers").exists():
            return candidate
        if (candidate / "papers.csv").exists() and (candidate / "literature").exists():
            return candidate
    return start.parent


ROOT = find_vault_root(PACKAGE_DIR)
PAPER_DIR = ROOT / "papers"
INDEX_PATH = ROOT / "papers.csv"
XLSX_PATH = ROOT / "paper_library.xlsx"
NOTES_DIR = ROOT / "notes"
EXTRACT_DIR = ROOT / "extracted_text"
FIGURE_DIR = ROOT / "figures"
OBSIDIAN_DIR = ROOT / "literature"
OBSIDIAN_PAPERS_DIR = OBSIDIAN_DIR / "papers"
OBSIDIAN_MAPS_DIR = OBSIDIAN_DIR / "maps"
OBSIDIAN_YEARS_DIR = OBSIDIAN_DIR / "years"
OBSIDIAN_VENUES_DIR = OBSIDIAN_DIR / "venues"
OBSIDIAN_TOPICS_DIR = OBSIDIAN_DIR / "topics"

READ_DONE = {"read", "done", "finished", "yes", "已读", "完成"}
NOTE_DONE = {"generated", "done", "yes", "已生成"}

CSV_FIELDS = [
    "paper_id",
    "filename",
    "title",
    "year",
    "venue",
    "subfield",
    "keywords",
    "importance",
    "importance_reason",
    "read_status",
    "note_status",
    "note_path",
    "selected_on",
    "last_reviewed_on",
    "added_on",
    "updated_on",
    "file_size_mb",
]

VENUES = {
    "AAAI",
    "ACL",
    "CVPR",
    "ECCV",
    "ICCV",
    "ICLR",
    "ICML",
    "IJCV",
    "MMASIA",
    "NEURIPS",
    "NIPS",
    "PMLR",
    "WACV",
}

LANDMARK_PATTERNS = {
    "denoising diffusion probabilistic models": "foundational diffusion model",
    "high-resolution image synthesis with latent diffusion models": "latent diffusion foundation",
    "taming transformers for high-resolution image synthesis": "VQGAN and transformer foundation",
    "maskgit": "masked visual token generation foundation",
    "make-a-video": "early large-scale text-to-video model",
    "cogvideo": "large text-to-video transformer line",
    "cogview": "large text-to-image transformer line",
    "diffusion models beat gans": "important diffusion scaling result",
    "scalable diffusion models with transformers": "DiT foundation",
    "visual autoregressive modeling": "major autoregressive image generation result",
    "wan open and advanced": "major open video generation system",
    "hunyuanvideo": "major video generation system",
    "open-sora": "open video generation system",
    "sana efficient high-resolution": "efficient high-resolution T2I system",
    "soundstream": "neural audio codec foundation",
}


def today() -> str:
    return dt.date.today().isoformat()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def slugify(text: str, limit: int = 80) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return text[:limit].strip("-") or "paper"


def parse_filename(path: Path) -> Dict[str, str]:
    stem = path.stem
    match = re.match(r"^(?P<year>\d{4})_(?P<rest>.+)$", stem)
    year = match.group("year") if match else ""
    rest = match.group("rest") if match else stem

    rest = rest.replace("_", " ")
    rest = normalize_space(rest)

    supplementary = bool(
        re.search(r"\b(supplementary material|supplemental|supplementary)\b", rest, re.I)
    )
    rest = re.sub(
        r"\b(supplementary material|supplemental|supplementary)\b",
        "",
        rest,
        flags=re.I,
    )
    rest = normalize_space(rest)

    venue = ""
    parts = rest.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].upper().replace("-", "") in VENUES:
        venue = parts[1].upper().replace("-", "")
        rest = parts[0].strip()

    if "ICLR Reject" in stem:
        venue = "ICLR Reject"

    title = normalize_space(rest)
    if supplementary:
        title = f"{title} (supplement)"

    return {
        "year": year,
        "title": title,
        "venue": venue,
        "is_supplement": "yes" if supplementary else "",
    }


def tag_title(title: str) -> Tuple[str, List[str]]:
    low = title.lower()
    tags: List[str] = []

    def add(name: str, *needles: str) -> None:
        if any(needle in low for needle in needles):
            tags.append(name)

    add("survey", "survey", "prospects")
    add("benchmark", "benchmark", "vbench")
    add("dataset", "dataset", "datacuration", "data curation", "captions")
    add("text-to-video", "text-to-video", "text to video", "t2v", "video generation", "video generative", "cogvideo")
    add("image-to-video", "image-to-video", "image to video", "i2v", "real image animation")
    add("long-video", "long video", "minute-long", "infinite videos", "longer video", "streamingt2v")
    add("text-to-image", "text-to-image", "text to image", "t2i", "image synthesis", "image generation", "image generative", "cogview")
    add("diffusion-flow", "diffusion", "rectified flow", "flow matching", "denoising")
    add("autoregressive", "autoregressive", "next-scale", "next-frame")
    add("visual-tokenizer-vq", "vector quant", "vq-", "vqgan", "vq-vae", "codebook", "tokenizer", "quantization")
    add("inpainting-completion", "inpainting", "completion")
    add("prompt-alignment", "prompt", "text understanding", "semantic fidelity", "attention regulation")
    add("vision-representation", "beit", "dinov2", "efficientvit", "masked autoencoders", "vision transformer", "navit")
    add("audio-speech", "audio", "soundstream", "speech")
    add("time-series", "time series", "forecaster")
    add("protein", "protein")

    if "survey" in tags:
        subfield = "Survey / Roadmap"
    elif "benchmark" in tags:
        subfield = "Benchmark / Evaluation"
    elif "dataset" in tags:
        subfield = "Dataset / Data Curation"
    elif "long-video" in tags:
        subfield = "Long Video Generation"
    elif "text-to-video" in tags or "image-to-video" in tags:
        subfield = "Video Generation"
    elif "inpainting-completion" in tags:
        subfield = "Image Inpainting / Completion"
    elif "visual-tokenizer-vq" in tags:
        subfield = "Visual Tokenization / Vector Quantization"
    elif "autoregressive" in tags:
        subfield = "Autoregressive Visual Generation"
    elif "diffusion-flow" in tags:
        subfield = "Diffusion / Flow Models"
    elif "text-to-image" in tags:
        subfield = "Text-to-Image Generation"
    elif "vision-representation" in tags:
        subfield = "Vision Representation / Architecture"
    elif "audio-speech" in tags:
        subfield = "Audio / Speech Generation"
    elif "time-series" in tags:
        subfield = "Time Series Modeling"
    elif "protein" in tags:
        subfield = "Protein / Biosequence Modeling"
    else:
        subfield = "Other Generative Modeling"

    return subfield, sorted(set(tags))


def infer_importance(title: str, year: str, venue: str, subfield: str, tags: Iterable[str]) -> Tuple[int, str]:
    low = title.lower()
    tag_set = set(tags)
    score = 3
    reasons: List[str] = []

    for pattern, reason in LANDMARK_PATTERNS.items():
        if pattern in low:
            score += 2
            reasons.append(reason)
            break

    if venue.upper() in {"CVPR", "ICCV", "ECCV", "ICLR", "ICML", "NEURIPS", "NIPS", "PMLR"}:
        score += 1
        reasons.append(f"top venue: {venue}")

    if tag_set & {"survey", "benchmark", "dataset"}:
        score += 1
        reasons.append("useful as overview/evaluation/data reference")

    try:
        if int(year) >= 2024 and subfield in {"Video Generation", "Long Video Generation"}:
            score += 1
            reasons.append("recent video-generation work")
    except ValueError:
        pass

    if "supplement" in low:
        score -= 2
        reasons.append("supplementary material")

    if "reject" in venue.lower():
        score -= 1
        reasons.append("marked as rejected version")

    if subfield in {"Time Series Modeling", "Protein / Biosequence Modeling"}:
        score -= 1
        reasons.append("peripheral to the main visual generation library")

    score = max(1, min(5, score))
    if not reasons:
        reasons.append("baseline relevance from filename heuristics")
    return score, "; ".join(reasons)


def load_index() -> Dict[str, Dict[str, str]]:
    if not INDEX_PATH.exists():
        return {}
    with INDEX_PATH.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = csv.DictReader(fh)
        return {row["paper_id"]: normalize_row(row) for row in rows if row.get("paper_id")}


def normalize_row(row: Dict[str, str]) -> Dict[str, str]:
    return {field: row.get(field, "") for field in CSV_FIELDS}


def resolve_library_path(path_text: str) -> Path:
    path = Path(path_text)
    return path if path.is_absolute() else ROOT / path


def library_relative_path(path_text: str) -> str:
    path = Path(path_text)
    full_path = path if path.is_absolute() else ROOT / path
    try:
        return full_path.resolve().relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return str(path_text).replace("\\", "/")


def write_index(rows: List[Dict[str, str]]) -> None:
    rows = [normalize_row(row) for row in rows]
    rows.sort(key=lambda r: (r.get("read_status") in READ_DONE, -safe_int(r.get("importance")), r.get("year", ""), r.get("title", "")))
    with INDEX_PATH.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def safe_int(value: Optional[str], default: int = 0) -> int:
    try:
        return int(str(value or "").strip())
    except ValueError:
        return default


def yaml_string(value: str) -> str:
    value = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{value}"'


def yaml_sequence(key: str, values: Iterable[str]) -> List[str]:
    cleaned = [str(value).strip() for value in values if str(value).strip()]
    if not cleaned:
        return [f"{key}: []"]
    return [f"{key}:"] + [f"  - {yaml_string(value)}" for value in cleaned]


def tag_slug(text: str) -> str:
    return slugify(text, limit=80).replace("-", "_")


def markdown_escape(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("$", "\\$").replace("\n", " ")


def wiki_target_escape(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("$", "\\$").replace("\n", " ")


def wiki_link(target: str, alias: str, *, table: bool = False) -> str:
    separator = "\\|" if table else "|"
    return f"[[{wiki_target_escape(target)}{separator}{markdown_escape(alias)}]]"


def note_wiki_link(row: Dict[str, str], *, table: bool = False) -> str:
    note_path = row.get("note_path", "")
    title = row.get("title", row.get("paper_id", "note"))
    if not note_path:
        return ""
    target = Path(library_relative_path(note_path)).with_suffix("").as_posix()
    return wiki_link(target, title, table=table)


def paper_wiki_link(row: Dict[str, str], *, table: bool = False) -> str:
    paper_id = row.get("paper_id", "")
    title = row.get("title") or paper_id or "paper"
    if not paper_id:
        return markdown_escape(title)
    return wiki_link(f"literature/papers/{paper_id}", title, table=table)


def pdf_wiki_link(row: Dict[str, str], *, table: bool = False) -> str:
    filename = row.get("filename", "")
    if not filename:
        return ""
    return wiki_link(f"papers/{filename}", "PDF", table=table)


def split_keywords(row: Dict[str, str]) -> List[str]:
    return [part.strip() for part in row.get("keywords", "").split(",") if part.strip()]


def field_wiki_link(subfield: str, *, table: bool = False) -> str:
    if not subfield:
        return ""
    return wiki_link(f"literature/fields/{slugify(subfield)}", subfield, table=table)


def year_wiki_link(year: str, *, table: bool = False) -> str:
    if not year:
        return ""
    return wiki_link(f"literature/years/{slugify(year)}", year, table=table)


def venue_wiki_link(venue: str, *, table: bool = False) -> str:
    if not venue:
        return ""
    return wiki_link(f"literature/venues/{slugify(venue)}", venue, table=table)


def topic_wiki_link(keyword: str, *, table: bool = False) -> str:
    if not keyword:
        return ""
    return wiki_link(f"literature/topics/{slugify(keyword)}", keyword, table=table)


def row_tags(row: Dict[str, str]) -> List[str]:
    tags = ["paper"]
    if row.get("subfield"):
        tags.append(f"paper/{tag_slug(row['subfield'])}")
    if row.get("venue"):
        tags.append(f"venue/{tag_slug(row['venue'])}")
    if row.get("year"):
        tags.append(f"year/{row['year']}")
    keywords = [part.strip() for part in row.get("keywords", "").split(",") if part.strip()]
    for keyword in keywords[:6]:
        tags.append(f"topic/{tag_slug(keyword)}")
    return sorted(set(tags))


def frontmatter_for_row(row: Dict[str, str]) -> str:
    tags = "\n".join(f"  - {tag}" for tag in row_tags(row))
    pdf_link = f"[[{wiki_target_escape('papers/' + row.get('filename', ''))}]]"
    paper_link = f"[[{wiki_target_escape('literature/papers/' + row.get('paper_id', ''))}]]"
    fields = [
        "---",
        "type: paper-note",
        *yaml_sequence("aliases", [row.get("title", "")]),
        f"paper_id: {yaml_string(row.get('paper_id', ''))}",
        f"title: {yaml_string(row.get('title', ''))}",
        f"year: {row.get('year') or 'null'}",
        f"venue: {yaml_string(row.get('venue', ''))}",
        f"subfield: {yaml_string(row.get('subfield', ''))}",
        *yaml_sequence("topics", split_keywords(row)),
        f"importance: {safe_int(row.get('importance'))}",
        f"read_status: {yaml_string(row.get('read_status', ''))}",
        f"note_status: {yaml_string(row.get('note_status', ''))}",
        f"selected_on: {yaml_string(row.get('selected_on', ''))}",
        f"last_reviewed_on: {yaml_string(row.get('last_reviewed_on', ''))}",
        f"paper: {yaml_string(paper_link)}",
        f"pdf: {yaml_string(pdf_link)}",
        "tags:",
        tags or "  - paper",
        "---",
        "",
    ]
    return "\n".join(fields)


def replace_frontmatter(content: str, frontmatter: str) -> str:
    body = content.lstrip("\ufeff")
    while body.startswith("---\n"):
        end = body.find("\n---\n", 4)
        if end == -1:
            break
        body = body[end + len("\n---\n") :].lstrip("\n")
    return frontmatter + body


def parse_frontmatter(content: str) -> Dict[str, str]:
    normalized = content.lstrip("\ufeff")
    if not normalized.startswith("---\n"):
        return {}
    end = normalized.find("\n---\n", 4)
    if end == -1:
        return {}
    metadata: Dict[str, str] = {}
    for line in normalized[4:end].splitlines():
        if not line or line.startswith(" ") or ":" not in line:
            continue
        key, value = line.split(":", 1)
        value = value.strip()
        if value == "null":
            value = ""
        if len(value) >= 2 and value[0] == value[-1] == '"':
            value = value[1:-1].replace('\\"', '"').replace("\\\\", "\\")
        metadata[key.strip()] = value
    return metadata


def apply_note_frontmatter_to_row(row: Dict[str, str]) -> bool:
    note_path = row.get("note_path", "")
    if not note_path:
        return False
    path = resolve_library_path(note_path)
    if not path.exists() or path.suffix.lower() != ".md":
        return False
    content = path.read_text(encoding="utf-8-sig")
    metadata = parse_frontmatter(content)
    if not metadata:
        return False

    changed = False
    editable_fields = ["read_status", "importance", "subfield", "venue", "selected_on", "last_reviewed_on"]
    for field in editable_fields:
        value = metadata.get(field)
        if value is None:
            continue
        if field == "importance":
            value = str(safe_int(value, safe_int(row.get(field))))
        if value != row.get(field, ""):
            row[field] = value
            changed = True

    if changed:
        if row.get("read_status", "").lower() in READ_DONE and not row.get("last_reviewed_on"):
            row["last_reviewed_on"] = today()
        row["updated_on"] = today()
    return changed


def sync_note_frontmatter(row: Dict[str, str]) -> bool:
    note_path = row.get("note_path", "")
    if not note_path:
        return False
    path = resolve_library_path(note_path)
    if not path.exists() or path.suffix.lower() != ".md":
        return False
    content = path.read_text(encoding="utf-8-sig")
    updated = replace_frontmatter(content, frontmatter_for_row(row))
    if updated != content:
        path.write_text(updated, encoding="utf-8")
    return True


def clean_extracted_text(text: str) -> str:
    text = text.replace("\x00", "")
    cleaned_lines = [normalize_space(line) for line in text.splitlines()]
    output: List[str] = []
    blank = False
    for line in cleaned_lines:
        if not line:
            if not blank:
                output.append("")
            blank = True
            continue
        output.append(line)
        blank = False
    return "\n".join(output).strip()


def scan(refresh_classification: bool = False) -> List[Dict[str, str]]:
    existing = load_index()
    now = today()
    rows: List[Dict[str, str]] = []
    seen_ids = set()

    search_dir = PAPER_DIR if PAPER_DIR.exists() else ROOT
    for path in sorted(search_dir.glob("*.pdf"), key=lambda p: p.name.lower()):
        parsed = parse_filename(path)
        paper_id = f"{parsed['year']}-{slugify(parsed['title'])}" if parsed["year"] else slugify(parsed["title"])
        original_id = paper_id
        counter = 2
        while paper_id in seen_ids:
            paper_id = f"{original_id}-{counter}"
            counter += 1
        seen_ids.add(paper_id)

        row = existing.get(paper_id, {}).copy()
        is_new = not row

        subfield, tags = tag_title(parsed["title"])
        importance, importance_reason = infer_importance(
            parsed["title"],
            parsed["year"],
            parsed["venue"],
            subfield,
            tags,
        )

        row.update(
            {
                "paper_id": paper_id,
                "filename": path.name,
                "title": parsed["title"],
                "year": parsed["year"],
                "venue": parsed["venue"],
                "file_size_mb": f"{path.stat().st_size / (1024 * 1024):.2f}",
                "updated_on": now,
            }
        )

        if is_new:
            row["added_on"] = now
            row["read_status"] = "unread"
            row["note_status"] = ""

        if refresh_classification or not row.get("subfield"):
            row["subfield"] = subfield
        if refresh_classification or not row.get("keywords"):
            row["keywords"] = ", ".join(tags)
        if refresh_classification or not row.get("importance"):
            row["importance"] = str(importance)
        if refresh_classification or not row.get("importance_reason"):
            row["importance_reason"] = importance_reason

        rows.append(normalize_row(row))

    write_index(rows)
    return rows


def unread_candidates(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    candidates = []
    for row in rows:
        read_status = row.get("read_status", "").strip().lower()
        note_status = row.get("note_status", "").strip().lower()
        if read_status in READ_DONE:
            continue
        if note_status in NOTE_DONE:
            continue
        candidates.append(row)
    return candidates


def pick(reserve: bool = False) -> Optional[Dict[str, str]]:
    rows = scan(refresh_classification=False)
    candidates = unread_candidates(rows)
    if not candidates:
        return None

    candidates.sort(
        key=lambda r: (
            safe_int(r.get("importance")),
            safe_int(r.get("year")),
            not bool(r.get("selected_on")),
            r.get("title", ""),
        ),
        reverse=True,
    )
    selected = candidates[0]

    if reserve:
        for row in rows:
            if row["paper_id"] == selected["paper_id"]:
                row["selected_on"] = today()
                row["note_status"] = "reserved"
                row["updated_on"] = today()
                selected = row
                break
        write_index(rows)

    return selected


def find_row(paper_id: str) -> Dict[str, str]:
    rows = scan(refresh_classification=False)
    for row in rows:
        if row["paper_id"] == paper_id:
            return row
    raise SystemExit(f"Paper id not found: {paper_id}")


def mark_note(paper_id: str, note_path: str) -> None:
    rows = scan(refresh_classification=False)
    found = False
    updated_row: Optional[Dict[str, str]] = None
    for row in rows:
        if row["paper_id"] == paper_id:
            row["note_status"] = "generated"
            row["note_path"] = library_relative_path(note_path)
            row["updated_on"] = today()
            found = True
            updated_row = row
            break
    if not found:
        raise SystemExit(f"Paper id not found: {paper_id}")
    write_index(rows)
    if updated_row:
        sync_note_frontmatter(updated_row)


def mark_read(paper_id: str, status: str = "read") -> None:
    rows = scan(refresh_classification=False)
    found = False
    updated_row: Optional[Dict[str, str]] = None
    for row in rows:
        if row["paper_id"] == paper_id:
            row["read_status"] = status
            row["last_reviewed_on"] = today()
            row["updated_on"] = today()
            found = True
            updated_row = row
            break
    if not found:
        raise SystemExit(f"Paper id not found: {paper_id}")
    write_index(rows)
    if updated_row:
        sync_note_frontmatter(updated_row)


def export_xlsx() -> None:
    rows = scan(refresh_classification=False)
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(f"Missing spreadsheet dependency: {exc}") from exc

    df = pd.DataFrame(rows, columns=CSV_FIELDS)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="papers")
        ws = writer.book["papers"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        widths = {
            "A": 34,
            "B": 70,
            "C": 70,
            "F": 36,
            "G": 42,
            "I": 56,
            "L": 42,
        }
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[letter].width = widths.get(letter, 16)


def extract_text(paper_id: str, max_chars: int, output: Optional[Path]) -> str:
    row = find_row(paper_id)
    path = PAPER_DIR / row["filename"]
    if not path.exists():
        path = ROOT / row["filename"]
    if not path.exists():
        raise SystemExit(f"PDF not found: {path}")
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise SystemExit(f"Missing PDF dependency: {exc}") from exc

    reader = PdfReader(str(path))
    chunks: List[str] = []
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:  # pragma: no cover - defensive for malformed PDFs
            page_text = f"[Page {page_number} extraction failed: {exc}]"
        if page_text.strip():
            chunks.append(f"\n\n--- Page {page_number} ---\n{page_text.strip()}")
        if sum(len(chunk) for chunk in chunks) >= max_chars:
            break

    text = clean_extracted_text("\n".join(chunks))
    text = text[:max_chars]
    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(text, encoding="utf-8")
    return text


def relative_to_root(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return path.resolve().as_posix()


def extract_images(
    paper_id: str,
    output_dir: Optional[Path],
    max_images: int,
    min_width: int,
    min_height: int,
) -> List[Dict[str, str]]:
    row = find_row(paper_id)
    path = PAPER_DIR / row["filename"]
    if not path.exists():
        path = ROOT / row["filename"]
    if not path.exists():
        raise SystemExit(f"PDF not found: {path}")
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise SystemExit(f"Missing PDF dependency: {exc}") from exc

    target_dir = output_dir or FIGURE_DIR / paper_id
    target_dir.mkdir(parents=True, exist_ok=True)
    reader = PdfReader(str(path))
    extracted: List[Dict[str, str]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        try:
            images = list(page.images)
        except Exception:
            images = []
        for image_index, image_file in enumerate(images, start=1):
            pil_image = getattr(image_file, "image", None)
            width = getattr(pil_image, "width", 0) or 0
            height = getattr(pil_image, "height", 0) or 0
            if width < min_width or height < min_height:
                continue

            source_name = getattr(image_file, "name", "") or f"image_{image_index}.png"
            suffix = Path(source_name).suffix.lower() or ".png"
            safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(source_name).stem) or "image"
            output_path = target_dir / f"page_{page_number:03d}_{image_index:02d}_{safe_name}{suffix}"

            data = getattr(image_file, "data", None)
            if data:
                output_path.write_bytes(data)
            elif pil_image is not None:
                pil_image.save(output_path)
            else:
                continue

            extracted.append(
                {
                    "path": relative_to_root(output_path),
                    "page": str(page_number),
                    "width": str(width),
                    "height": str(height),
                    "source_name": source_name,
                }
            )
            if len(extracted) >= max_images:
                metadata_path = target_dir / "images.json"
                metadata_path.write_text(
                    json.dumps(extracted, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                return extracted

    metadata_path = target_dir / "images.json"
    metadata_path.write_text(
        json.dumps(extracted, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return extracted


def sort_rows_for_reading(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        rows,
        key=lambda r: (
            r.get("read_status", "").lower() in READ_DONE,
            -safe_int(r.get("importance")),
            -safe_int(r.get("year")),
            r.get("title", ""),
        ),
    )


def markdown_table(rows: List[Dict[str, str]], include_note: bool = True) -> str:
    if include_note:
        header = "| 论文 | 年份 | 领域 | 重要度 | 状态 | 笔记 | PDF |"
        divider = "|---|---:|---|---:|---|---|---|"
    else:
        header = "| 论文 | 年份 | 领域 | 重要度 | 状态 | PDF |"
        divider = "|---|---:|---|---:|---|---|"
    lines = [header, divider]
    for row in rows:
        title = paper_wiki_link(row, table=True)
        year = year_wiki_link(row.get("year", ""), table=True)
        subfield = field_wiki_link(row.get("subfield", ""), table=True)
        importance = markdown_escape(row.get("importance", ""))
        status = markdown_escape(row.get("read_status", ""))
        pdf = pdf_wiki_link(row, table=True)
        if include_note:
            note = note_wiki_link(row, table=True)
            lines.append(f"| {title} | {year} | {subfield} | {importance} | {status} | {note} | {pdf} |")
        else:
            lines.append(f"| {title} | {year} | {subfield} | {importance} | {status} | {pdf} |")
    return "\n".join(lines)


def write_markdown(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.rstrip() + "\n", encoding="utf-8")


def literature_frontmatter(page_type: str) -> str:
    return "\n".join(
        [
            "---",
            f"type: {page_type}",
            f"generated_on: {today()}",
            "tags:",
            "  - paper-library",
            "---",
            "",
        ]
    )


def paper_frontmatter(row: Dict[str, str]) -> str:
    tags = "\n".join(f"  - {tag}" for tag in row_tags(row))
    pdf_link = f"[[{wiki_target_escape('papers/' + row.get('filename', ''))}]]"
    fields = [
        "---",
        "type: paper",
        *yaml_sequence("aliases", [row.get("title", "")]),
        f"paper_id: {yaml_string(row.get('paper_id', ''))}",
        f"title: {yaml_string(row.get('title', ''))}",
        f"year: {row.get('year') or 'null'}",
        f"venue: {yaml_string(row.get('venue', ''))}",
        f"subfield: {yaml_string(row.get('subfield', ''))}",
        *yaml_sequence("topics", split_keywords(row)),
        f"importance: {safe_int(row.get('importance'))}",
        f"read_status: {yaml_string(row.get('read_status', ''))}",
        f"note_status: {yaml_string(row.get('note_status', ''))}",
        f"added_on: {yaml_string(row.get('added_on', ''))}",
        f"updated_on: {yaml_string(row.get('updated_on', ''))}",
        f"pdf: {yaml_string(pdf_link)}",
        "tags:",
        tags or "  - paper",
        "---",
        "",
    ]
    return "\n".join(fields)


def topic_frequency(rows: List[Dict[str, str]]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        for keyword in split_keywords(row):
            counts[keyword] = counts.get(keyword, 0) + 1
    return counts


def graph_topic_weight(keyword: str, counts: Dict[str, int]) -> int:
    count = counts.get(keyword, 0)
    if count <= 3:
        return 5
    if count <= 8:
        return 4
    if count <= 20:
        return 2
    return 1


def related_papers(row: Dict[str, str], rows: List[Dict[str, str]], limit: int = 4) -> List[Dict[str, str]]:
    row_keywords = set(split_keywords(row))
    keyword_counts = topic_frequency(rows)
    ignored_terms = {
        "with",
        "from",
        "using",
        "based",
        "model",
        "models",
        "generation",
        "generative",
        "image",
        "video",
        "text",
        "towards",
        "learning",
        "large",
        "scale",
        "open",
        "paper",
    }
    row_title_terms = {
        term
        for term in re.findall(r"[a-z0-9]+", row.get("title", "").lower())
        if len(term) >= 4 and term not in ignored_terms
    }

    scored: List[Tuple[int, int, int, str, Dict[str, str]]] = []
    for candidate in rows:
        if candidate.get("paper_id") == row.get("paper_id"):
            continue

        candidate_keywords = set(split_keywords(candidate))
        candidate_title_terms = {
            term
            for term in re.findall(r"[a-z0-9]+", candidate.get("title", "").lower())
            if len(term) >= 4 and term not in ignored_terms
        }
        score = 0
        if row.get("subfield") and row.get("subfield") == candidate.get("subfield"):
            score += 2
        shared_keywords = row_keywords & candidate_keywords
        score += sum(graph_topic_weight(keyword, keyword_counts) for keyword in shared_keywords)
        score += min(len(row_title_terms & candidate_title_terms), 3)
        if score < 3:
            continue

        scored.append(
            (
                score,
                safe_int(candidate.get("importance")),
                safe_int(candidate.get("year")),
                candidate.get("title", ""),
                candidate,
            )
        )

    scored.sort(key=lambda item: (-item[0], -item[1], -item[2], item[3]))
    return [item[-1] for item in scored[:limit]]


def paper_page_content(row: Dict[str, str], rows: List[Dict[str, str]]) -> str:
    keywords = split_keywords(row)
    topic_links = ", ".join(topic_wiki_link(keyword) for keyword in keywords) or "-"
    note = note_wiki_link(row) or "-"
    venue = venue_wiki_link(row.get("venue", "")) or "-"
    related_lines = [
        f"- {paper_wiki_link(candidate)} ({candidate.get('year', '')}, {candidate.get('venue', '') or 'no venue'})"
        for candidate in related_papers(row, rows)
    ]
    if not related_lines:
        related_lines = ["- No close local match yet."]

    return (
        paper_frontmatter(row)
        + f"# {markdown_escape(row.get('title', row.get('paper_id', 'Paper')))}\n\n"
        + "## Metadata\n\n"
        + f"- Field: {field_wiki_link(row.get('subfield', '')) or '-'}\n"
        + f"- Year: {year_wiki_link(row.get('year', '')) or '-'}\n"
        + f"- Venue: {venue}\n"
        + f"- Topics: {topic_links}\n"
        + f"- Importance: {row.get('importance', '') or '-'} / 5\n"
        + f"- Status: {row.get('read_status', '') or '-'}\n"
        + f"- Note: {note}\n"
        + f"- PDF: {pdf_wiki_link(row) or '-'}\n"
        + f"- Added: {row.get('added_on', '') or '-'}\n"
        + f"- Updated: {row.get('updated_on', '') or '-'}\n\n"
        + "## Why It Is Here\n\n"
        + f"{row.get('importance_reason', '') or 'No local rationale recorded yet.'}\n\n"
        + "## Related Papers\n\n"
        + "\n".join(related_lines)
        + "\n"
    )


def write_group_index(
    directory: Path,
    key: str,
    rows: List[Dict[str, str]],
    page_type: str,
    title: str,
    extra_links: Optional[List[str]] = None,
) -> None:
    extra = ""
    if extra_links:
        extra = "\n".join(f"- {link}" for link in extra_links if link) + "\n\n"
    content = (
        literature_frontmatter(page_type)
        + f"# {title}\n\n"
        + extra
        + f"- Papers: {len(rows)}\n\n"
        + markdown_table(sort_rows_for_reading(rows))
    )
    write_markdown(directory / f"{slugify(key)}.md", content)


def graph_paper_line(row: Dict[str, str]) -> str:
    topics = ", ".join(topic_wiki_link(keyword) for keyword in split_keywords(row)[:3]) or "-"
    venue = row.get("venue") or "no venue"
    return (
        f"- {paper_wiki_link(row)} "
        f"({row.get('year', '') or 'no year'}, {venue}; "
        f"{field_wiki_link(row.get('subfield', '')) or '-'}; {topics})"
    )


def graph_overview_content(
    rows: List[Dict[str, str]],
    by_subfield: Dict[str, List[Dict[str, str]]],
    by_topic: Dict[str, List[Dict[str, str]]],
) -> str:
    fields = sorted(by_subfield.items(), key=lambda item: (-len(item[1]), item[0]))
    topics = sorted(by_topic.items(), key=lambda item: (-len(item[1]), item[0]))[:18]
    anchors = sort_rows_for_reading([row for row in rows if safe_int(row.get("importance")) >= 5])[:18]

    field_lines = [
        f"- {field_wiki_link(field)} ({len(group)} 篇)"
        for field, group in fields
    ]
    topic_lines = [
        f"- {topic_wiki_link(topic)} ({len(group)} 篇)"
        for topic, group in topics
    ]
    anchor_lines = [graph_paper_line(row) for row in anchors]

    return (
        literature_frontmatter("paper-graph-map")
        + "# 论文知识图谱总览\n\n"
        + "这页是全局关系图谱的主入口：领域负责分层，主题负责横向连接，重点论文负责把图谱锚定到具体工作。\n\n"
        + "## 主干领域\n\n"
        + "\n".join(field_lines)
        + "\n\n## 主题轴\n\n"
        + "\n".join(topic_lines)
        + "\n\n## 核心论文锚点\n\n"
        + "\n".join(anchor_lines)
        + "\n\n## 其他图谱页\n\n"
        + "- [[literature/maps/topic-bridges|跨主题桥接]]\n"
        + "- [[literature/maps/research-frontier|重点阅读路线]]\n"
    )


def graph_topic_bridge_content(
    rows: List[Dict[str, str]],
    by_topic: Dict[str, List[Dict[str, str]]],
) -> str:
    top_topics = sorted(by_topic.items(), key=lambda item: (-len(item[1]), item[0]))[:24]
    topic_lines: List[str] = []
    for topic, group in top_topics:
        fields = sorted({row.get("subfield", "") for row in group if row.get("subfield")})
        field_links = ", ".join(field_wiki_link(field) for field in fields[:5]) or "-"
        topic_lines.append(f"- {topic_wiki_link(topic)} -> {field_links}")

    bridge_rows = [
        row
        for row in rows
        if safe_int(row.get("importance")) >= 4 and len(split_keywords(row)) >= 2
    ]
    bridge_lines = [graph_paper_line(row) for row in sort_rows_for_reading(bridge_rows)[:28]]

    return (
        literature_frontmatter("paper-graph-map")
        + "# 跨主题桥接\n\n"
        + "这页只保留能跨领域或跨方法连接的主题和论文，避免全局图被年份、会场和普通列表边淹没。\n\n"
        + "## 主题到领域\n\n"
        + "\n".join(topic_lines)
        + "\n\n## 桥接论文\n\n"
        + ("\n".join(bridge_lines) if bridge_lines else "- No bridge papers yet.")
        + "\n"
    )


def graph_frontier_content(
    by_subfield: Dict[str, List[Dict[str, str]]],
) -> str:
    sections: List[str] = []
    for subfield, group in sorted(by_subfield.items()):
        priority = [
            row
            for row in sort_rows_for_reading(group)
            if safe_int(row.get("importance")) >= 4
        ][:6]
        if not priority:
            continue
        sections.append(f"## {subfield}\n")
        sections.extend(graph_paper_line(row) for row in priority)
        sections.append("")

    return (
        literature_frontmatter("paper-graph-map")
        + "# 重点阅读路线\n\n"
        + "这页按领域挑选高重要度论文，适合在图谱中从领域节点跳到代表性论文，再沿相关论文继续展开。\n\n"
        + "\n".join(sections).rstrip()
        + "\n"
    )


def generate_obsidian_pages() -> Dict[str, int]:
    rows = scan(refresh_classification=False)
    OBSIDIAN_DIR.mkdir(parents=True, exist_ok=True)
    fields_dir = OBSIDIAN_DIR / "fields"
    fields_dir.mkdir(parents=True, exist_ok=True)
    for directory in (
        OBSIDIAN_PAPERS_DIR,
        OBSIDIAN_MAPS_DIR,
        OBSIDIAN_YEARS_DIR,
        OBSIDIAN_VENUES_DIR,
        OBSIDIAN_TOPICS_DIR,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    imported_notes = 0
    for row in rows:
        if apply_note_frontmatter_to_row(row):
            imported_notes += 1
    if imported_notes:
        write_index(rows)
        rows = scan(refresh_classification=False)

    synced_notes = 0
    for row in rows:
        if sync_note_frontmatter(row):
            synced_notes += 1

    unread = [row for row in rows if row.get("read_status", "").lower() not in READ_DONE]
    read = [row for row in rows if row.get("read_status", "").lower() in READ_DONE]
    noted = [row for row in rows if row.get("note_status", "").lower() in NOTE_DONE]
    high_priority = [row for row in rows if safe_int(row.get("importance")) >= 5]

    by_subfield: Dict[str, List[Dict[str, str]]] = {}
    by_year: Dict[str, List[Dict[str, str]]] = {}
    by_venue: Dict[str, List[Dict[str, str]]] = {}
    by_topic: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        subfield = row.get("subfield") or "Uncategorized"
        by_subfield.setdefault(subfield, []).append(row)
        if row.get("year"):
            by_year.setdefault(row["year"], []).append(row)
        if row.get("venue"):
            by_venue.setdefault(row["venue"], []).append(row)
        for keyword in split_keywords(row):
            by_topic.setdefault(keyword, []).append(row)

    write_group_index(OBSIDIAN_PAPERS_DIR, "index", rows, "paper-index", "All Papers")
    for row in rows:
        write_markdown(OBSIDIAN_PAPERS_DIR / f"{row['paper_id']}.md", paper_page_content(row, rows))
    for year, group in sorted(by_year.items(), reverse=True):
        write_group_index(OBSIDIAN_YEARS_DIR, year, group, "paper-year-index", year)
    for venue, group in sorted(by_venue.items()):
        write_group_index(OBSIDIAN_VENUES_DIR, venue, group, "paper-venue-index", venue)
    for topic, group in sorted(by_topic.items()):
        fields = sorted({row.get("subfield", "") for row in group if row.get("subfield")})
        extra_links = [field_wiki_link(field) for field in fields[:8]]
        write_group_index(OBSIDIAN_TOPICS_DIR, topic, group, "paper-topic-index", topic, extra_links)

    write_markdown(OBSIDIAN_MAPS_DIR / "overview.md", graph_overview_content(rows, by_subfield, by_topic))
    write_markdown(OBSIDIAN_MAPS_DIR / "topic-bridges.md", graph_topic_bridge_content(rows, by_topic))
    write_markdown(OBSIDIAN_MAPS_DIR / "research-frontier.md", graph_frontier_content(by_subfield))

    subfield_lines = ["| 领域 | 篇数 | 未读 | 已读 |", "|---|---:|---:|---:|"]
    for subfield, group in sorted(by_subfield.items()):
        field_read = sum(1 for row in group if row.get("read_status", "").lower() in READ_DONE)
        field_unread = len(group) - field_read
        slug = slugify(subfield)
        field_topics = sorted({keyword for row in group for keyword in split_keywords(row)})
        field_topic_line = ", ".join(topic_wiki_link(keyword) for keyword in field_topics[:12]) or "-"
        subfield_lines.append(
            f"| {field_wiki_link(subfield, table=True)} | {len(group)} | {field_unread} | {field_read} |"
        )

        field_page = (
            literature_frontmatter("paper-field-index")
            + f"# {subfield}\n\n"
            + f"- 总数: {len(group)}\n"
            + f"- 未读: {field_unread}\n"
            + f"- 已读: {field_read}\n\n"
            + f"- Topics: {field_topic_line}\n\n"
            + markdown_table(sort_rows_for_reading(group))
        )
        write_markdown(fields_dir / f"{slug}.md", field_page)

    index_content = (
        literature_frontmatter("paper-library-index")
        + "# 论文库\n\n"
        + f"- 总论文数: {len(rows)}\n"
        + f"- 未读: {len(unread)}\n"
        + f"- 已读: {len(read)}\n"
        + f"- 已生成精讲笔记: {len(noted)}\n\n"
        + "## 快速入口\n\n"
        + "- [[literature/unread|未读论文]]\n"
        + "- [[literature/read|已读论文]]\n"
        + "- [[literature/high-priority|高优先级论文]]\n"
        + "- [[literature/notes|已生成精讲笔记]]\n\n"
        + "## 知识图谱入口\n\n"
        + "- [[literature/maps/overview|论文知识图谱总览]]\n"
        + "- [[literature/maps/topic-bridges|跨主题桥接]]\n"
        + "- [[literature/maps/research-frontier|重点阅读路线]]\n\n"
        + "## 细分领域\n\n"
        + "\n".join(subfield_lines)
        + "\n\n## 最近生成的精讲笔记\n\n"
        + markdown_table(sort_rows_for_reading(noted)[:12])
        + "\n\n## 检索索引\n\n"
        + "- [[literature/papers/index|全部论文页]]\n"
        + "- [[literature/years/2025|2025 论文]]\n"
        + "- [[literature/venues/cvpr|CVPR 论文]]\n"
        + "- [[literature/topics/text-to-video|text-to-video 论文]]\n"
    )
    write_markdown(OBSIDIAN_DIR / "index.md", index_content)

    unread_content = (
        literature_frontmatter("paper-reading-list")
        + "# 未读论文\n\n"
        + "按重要程度和年份排序。\n\n"
        + markdown_table(sort_rows_for_reading(unread))
    )
    write_markdown(OBSIDIAN_DIR / "unread.md", unread_content)

    read_content = (
        literature_frontmatter("paper-reading-list")
        + "# 已读论文\n\n"
        + markdown_table(sort_rows_for_reading(read))
    )
    write_markdown(OBSIDIAN_DIR / "read.md", read_content)

    high_priority_content = (
        literature_frontmatter("paper-reading-list")
        + "# 高优先级论文\n\n"
        + "包含 `importance >= 5` 的论文。\n\n"
        + markdown_table(sort_rows_for_reading(high_priority))
    )
    write_markdown(OBSIDIAN_DIR / "high-priority.md", high_priority_content)

    notes_content = (
        literature_frontmatter("paper-note-index")
        + "# 已生成精讲笔记\n\n"
        + markdown_table(sort_rows_for_reading(noted))
    )
    write_markdown(OBSIDIAN_DIR / "notes.md", notes_content)

    dataview_content = (
        literature_frontmatter("paper-library-help")
        + "# Dataview 查询示例\n\n"
        + "如果 Obsidian 安装了 Dataview 插件，可以使用下面的查询。\n\n"
        + "## 未读论文笔记\n\n"
        + "```dataview\n"
        + "TABLE year, venue, subfield, importance, pdf\n"
        + "FROM \"literature/papers\"\n"
        + "WHERE type = \"paper\" AND read_status != \"read\"\n"
        + "SORT importance DESC, year DESC\n"
        + "```\n\n"
        + "## 视频生成方向\n\n"
        + "```dataview\n"
        + "TABLE year, venue, importance, read_status\n"
        + "FROM \"literature/papers\"\n"
        + "WHERE type = \"paper\" AND subfield = \"Video Generation\"\n"
        + "SORT year DESC\n"
        + "```\n"
    )
    write_markdown(OBSIDIAN_DIR / "dataview.md", dataview_content)

    return {
        "papers": len(rows),
        "unread": len(unread),
        "read": len(read),
        "notes": len(noted),
        "subfields": len(by_subfield),
        "maps": 3,
        "paper_pages": len(rows),
        "years": len(by_year),
        "venues": len(by_venue),
        "topics": len(by_topic),
        "imported_notes": imported_notes,
        "synced_notes": synced_notes,
    }


def stats() -> Dict[str, Dict[str, int]]:
    rows = scan(refresh_classification=False)
    result: Dict[str, Dict[str, int]] = {
        "read_status": {},
        "note_status": {},
        "subfield": {},
        "importance": {},
    }
    for row in rows:
        for key in result:
            value = row.get(key) or "(blank)"
            result[key][value] = result[key].get(value, 0) + 1
    return result


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Manage local PDF paper library.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="scan PDFs and update papers.csv")
    scan_parser.add_argument("--refresh-classification", action="store_true")

    pick_parser = subparsers.add_parser("pick", help="pick one unread paper for a note")
    pick_parser.add_argument("--reserve", action="store_true", help="mark selected paper as reserved")

    extract_parser = subparsers.add_parser("extract", help="extract text from a paper PDF")
    extract_parser.add_argument("--paper-id", required=True)
    extract_parser.add_argument("--max-chars", type=int, default=45000)
    extract_parser.add_argument("--output", type=Path)

    image_parser = subparsers.add_parser("extract-images", help="extract embedded images from a paper PDF")
    image_parser.add_argument("--paper-id", required=True)
    image_parser.add_argument("--output-dir", type=Path)
    image_parser.add_argument("--max-images", type=int, default=24)
    image_parser.add_argument("--min-width", type=int, default=180)
    image_parser.add_argument("--min-height", type=int, default=120)

    note_parser = subparsers.add_parser("mark-note", help="mark a note as generated")
    note_parser.add_argument("--paper-id", required=True)
    note_parser.add_argument("--note-path", required=True)

    read_parser = subparsers.add_parser("mark-read", help="mark a paper read")
    read_parser.add_argument("--paper-id", required=True)
    read_parser.add_argument("--status", default="read")

    subparsers.add_parser("export-xlsx", help="export paper_library.xlsx")
    subparsers.add_parser("sync-obsidian", help="generate Obsidian index pages and note frontmatter")
    subparsers.add_parser("stats", help="print library stats")

    args = parser.parse_args(argv)

    if args.command == "scan":
        rows = scan(refresh_classification=args.refresh_classification)
        print(f"Indexed {len(rows)} PDFs into {INDEX_PATH.name}")
    elif args.command == "pick":
        selected = pick(reserve=args.reserve)
        if not selected:
            print("{}")
        else:
            print(json.dumps(selected, ensure_ascii=False, indent=2))
    elif args.command == "extract":
        text = extract_text(args.paper_id, args.max_chars, args.output)
        if args.output:
            print(f"Wrote {len(text)} chars to {args.output}")
        else:
            print(text)
    elif args.command == "extract-images":
        images = extract_images(
            args.paper_id,
            args.output_dir,
            args.max_images,
            args.min_width,
            args.min_height,
        )
        print(json.dumps(images, ensure_ascii=False, indent=2))
    elif args.command == "mark-note":
        mark_note(args.paper_id, args.note_path)
        print(f"Marked note generated for {args.paper_id}")
    elif args.command == "mark-read":
        mark_read(args.paper_id, args.status)
        print(f"Marked {args.paper_id} as {args.status}")
    elif args.command == "export-xlsx":
        export_xlsx()
        print(f"Exported {XLSX_PATH.name}")
    elif args.command == "sync-obsidian":
        print(json.dumps(generate_obsidian_pages(), ensure_ascii=False, indent=2))
    elif args.command == "stats":
        print(json.dumps(stats(), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
